from pathlib import Path

from openpyxl import load_workbook

from t2r.infra.export.xlsx import write_xlsx


def test_writes_headers_and_rows(tmp_path: Path):
    path = tmp_path / "out.xlsx"
    write_xlsx(
        str(path),
        columns=["id", "name"],
        rows=[[1, "a"], [2, "b"]],
        title="hello",
    )
    assert path.exists()
    wb = load_workbook(path)
    ws = wb.active
    assert [c.value for c in ws[1]] == ["id", "name"]
    assert [c.value for c in ws[2]] == [1, "a"]
    assert ws.freeze_panes == "A2"


def test_handles_complex_cells(tmp_path: Path):
    path = tmp_path / "complex.xlsx"
    write_xlsx(
        str(path),
        columns=["a", "b"],
        rows=[[None, {"k": "v"}], [[1, 2], 3]],
    )
    wb = load_workbook(path)
    ws = wb.active
    # complex values get serialized via str()
    assert ws.cell(row=2, column=2).value == "{'k': 'v'}"
    assert ws.cell(row=3, column=1).value == "[1, 2]"
