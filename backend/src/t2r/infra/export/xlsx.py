from __future__ import annotations

import os
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_xlsx(
    path: str,
    *,
    columns: list[str],
    rows: Iterable[list[Any]],
    title: str | None = None,
) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = (title or "Report")[:31]

    # Header
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FFC2410C")
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Rows
    for ri, row in enumerate(rows, start=2):
        for ci, value in enumerate(row, start=1):
            if isinstance(value, (dict, list)):
                value = str(value)
            ws.cell(row=ri, column=ci, value=value)

    # Auto-size columns to a reasonable max
    for ci in range(1, len(columns) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(path)
